import PyPDF2
import openpyxl
import pdfplumber
import pandas as pd
import re


# Set the path to your PDF file
pdf_path1 =  r"export_supplier_2023-09-25_POE9G93589228S_V1 noon.pdf"
# Initialize an empty list to store the extracted tables
tables = []

# Initialize pdfplumber object
with pdfplumber.open(pdf_path1) as pdf:
    # Iterate over each page and extract the tables
    for page in pdf.pages:
        # Extract tables from the current page
        page_tables = page.extract_tables()
        # Append the extracted tables to the list
        tables.extend(page_tables)

# Combine the extracted tables into a single DataFrame
combined_table = pd.concat([pd.DataFrame(table) for table in tables], ignore_index=True)

# Set the path to save the CSV file
csv_path1 =  r"noonoutput.xlsx"

# Save the DataFrame as an Excel file (CSV format)
combined_table.to_excel(csv_path1, sheet_name='Sheet1', index=False, header=False)

df = pd.read_excel(csv_path1)

# Create a new DataFrame with non-null values
final_list = []
for index, row in df.iterrows():
    roww = []
    for x in range(len(df.columns)):
        
        if  (x==0 or x==1 or x==3 or x==2 or x==4) and pd.isnull(row[x]) and row[x]!="SR No" and row[x]!="None":
            pass
        else:
            try:
                roww.append(row[x].replace("None","---"))
            except:
                roww.append(row[x])
    final_list.append(roww)

newdf = pd.DataFrame(final_list)

newdf.to_excel(csv_path1,index=False)

def delete_rows_before_keyword(excel_path, sheet_name, keyword):
    # Open the Excel file
    wb = openpyxl.load_workbook(excel_path)

    # Select the sheet by name
    sheet = wb[sheet_name]

    # Find the row index where the keyword is present
    target_row_index = None
    for row_idx, row in enumerate(sheet.iter_rows(min_col=1), start=1):
        if row[0].value == keyword or row[1].value == keyword:
            target_row_index = row_idx
            break

    if target_row_index is not None:
        # Delete rows before and including the target row
        sheet.delete_rows(1, target_row_index)

        # Save the modified Excel file
        wb.save(excel_path)
        print("Rows deleted successfully.")
    else:
        print("Keyword not found in the sheet.")

# Example usage
sheet_name = 'Sheet1'
keyword = 'SR No'
delete_rows_before_keyword(csv_path1, sheet_name, keyword)


def delete_rows_after_subtotal(excel_path, sheet_name):
    # Open the Excel file
    wb = openpyxl.load_workbook(excel_path)

    # Select the sheet by name
    sheet = wb[sheet_name]

    # Find the row index where the first cell contains "SUBTOTAL"
    subtotal_row_index = None
    for row in sheet.iter_rows():
        if row[0].value == "Sub Total (SAR)" or row[1].value == "Sub Total (SAR)":
            subtotal_row_index = row[0].row
            break

    # Delete rows after the subtotal row if found
    if subtotal_row_index:
        last_row_index = sheet.max_row
        if subtotal_row_index < last_row_index:
            sheet.delete_rows(subtotal_row_index, last_row_index)

    # Save the modified Excel file
    wb.save(excel_path)

    # Load the modified Excel file into a DataFrame
    df_modified = pd.read_excel(excel_path)

    # Return the modified DataFrame
    return df_modified

# Set the path to your Excel file

# Specify the sheet name
sheet_name = "Sheet1"

# Call the function
modified_df_2 = delete_rows_after_subtotal(csv_path1, sheet_name)

def extract_purchase_order_details(pdf_path):
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)

        page = reader.pages[0]
        text = page.extract_text()

        # Extract Purchase Order Number
        po_number = re.search(r'P\.O\s*No\s*:\s*([^\n]+)', text)
        po_str = po_number.group(1).strip() if po_number else None

        # Extract Ship To Location
        ship_to_location = re.search(r'Ship\s*To\s*:\s*([^\n]+)', text)
        ship_to_location_str = ship_to_location.group(1).strip() if ship_to_location else None

        # Extract Approval Date
        approval_date = re.search(r'Approval\s*Date:\s*([-0-9A-Za-z/]+)', text)
        approval_date_str = approval_date.group(1) if approval_date else None

        return po_str, ship_to_location_str, approval_date_str


po_str, shipto, aprrovalDate = extract_purchase_order_details(pdf_path1)

# Load the Excel file
excel_file = pd.ExcelFile(csv_path1)

# Create a new sheet for PO info
with pd.ExcelWriter(csv_path1, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
    # Write the purchase order details to the new sheet
    po_info_df = pd.DataFrame({
        'Purchase Order Number': [po_str],
        'Ship to': [shipto],
        'Approval Date': [aprrovalDate],

    })
    po_info_df.to_excel(writer, sheet_name='po info', index=False)


print("PO No: " + (po_str or "Not Found"))
print("Ship to: " + (shipto or "Not Found"))
print("Approval Date: " + (aprrovalDate or "Not Found"))


def update_xlsx_file(filename):
    # Load the workbook
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook["Sheet1"]  # Assuming the sheet name is "Sheet1"

    # Find the maximum row number in the sheet
    max_row = sheet.max_row

    # Iterate through each row in the sheet
    for row_num in range(1, max_row + 1):
        cell_value = sheet.cell(row=row_num, column=13).value

        if cell_value is None:  # Check if the cell is empty
            # Shift data to the right by one position starting from column 8
            for col_num in range(13, 7, -1):
                sheet.cell(row=row_num, column=col_num + 1).value = sheet.cell(row=row_num, column=col_num).value

            sheet.cell(row=row_num, column=8).value = None  # Set column 8 to empty

    # Save the changes
    workbook.save(filename)


# Call the function and provide the filename of the xlsx file
update_xlsx_file(csv_path1)