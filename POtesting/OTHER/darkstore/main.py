import PyPDF2
import openpyxl
import pdfplumber
import pandas as pd
import re
import numpy as np


# Set the path to your PDF file
pdf_path1 = r"Basamh (Goody) - باسمح-Starlink DC - RYD-PO1221126.pdf"
#pdf_path1 ="Basamh Commercial Distribution Company شركة باسمح للتوزيع التجاري-Riyadh - Qairawan Dark Store-PO879675.pdf"
# Initialize an empty list to store the extracted tables
tables = []

# Initialize pdfplumber object
with pdfplumber.open(pdf_path1) as pdf:
    # Iterate over each page and extract the tables
    for page in pdf.pages:
        # Extract tables from the current page
        page_tables = page.extract_tables({"horizontal_strategy":"text"})
        # Append the extracted tables to the list
        tables.extend(page_tables)

# Combine the extracted tables into a single DataFrame
combined_table = pd.concat([pd.DataFrame(table) for table in tables], ignore_index=True)

# Set the path to save the CSV file
csv_path1 = r"darkstoreoutput.xlsx"
#csv_path1 = r"darkstoreoutput.xlsx"
# Save the DataFrame as an Excel file (CSV format)
combined_table.to_excel(csv_path1, sheet_name='Original', index=False, header=False)

def shift_cells(xlsx_file):
    # Load the workbook
    wb = openpyxl.load_workbook(xlsx_file)

    # Select the first sheet
    sheet = wb.active

    # Convert the rows generator object to a list and iterate over it in reverse order
    rows = list(sheet.iter_rows(min_row=1, max_row=sheet.max_row))
    for row in reversed(rows):
        # Check if the first cell in the row is empty
        if row[0].value is None:
            # Shift cells to the left by one position
            for cell in row:
                if cell.column == 1:
                    cell.value = None
                else:
                    cell.value = sheet.cell(row=cell.row, column=cell.column+1).value

    # Remove empty rows
    for row in reversed(rows):
        if all(cell.value is None for cell in row):
            sheet.delete_rows(row[0].row)

    # Save the modified workbook
    wb.save(xlsx_file)

    print("Cell shifting and empty row removal complete!")
# Usage example:
shift_cells(csv_path1)
#### Mapping function
def mapcustomer_sku(customer_SKU,vendor,masterdatadf):
    master_data_file=r"C:\Users\rpauser\Documents\Item Master Date.xlsx"
    customer_SKU=customer_SKU
    mapped_BTC_code=""
    division_factor=""
    FOUND="FALSE"
    sheet_name = vendor


    xls = masterdatadf
    

    dfs = {sheet_name: xls.parse(sheet_name) for sheet_name in ["Dark Store"]}


    if sheet_name in dfs and customer_SKU!="":
        sheet_df = dfs[sheet_name]
        #print(sheet_df.iloc[:,0])
        try:
            result = sheet_df[sheet_df.iloc[:,0] == int(customer_SKU)]
        except:
            result = sheet_df[sheet_df.iloc[:,0] == customer_SKU]
        
        if not result.empty:
            mapped_BTC_code = int(result.iloc[0][1])
            if np.isnan(result.iloc[0][2]):
                division_factor = ""
            else:
                division_factor = result.iloc[0][2]
            FOUND="TRUE"
            
            #print(str(mapped_BTC_code),division_factor)
        else:
            pass
            #print("not found on {sheet_name}.")
    else:
        pass
        #print(f"{sheet_name} not found in the Excel file.")

    '''
    RPAEngine.SetVar("",mapped_BTC_code)
    RPAEngine.SetVar("",division_factor)
    RPAEngine.SetVar("",FOUND)
    '''
    return (mapped_BTC_code,division_factor,FOUND)

####
def add_mappeddata_toDF(xlsx_file):
        
    # Replace 'your_excel_file.xlsx' with the path to your Excel file
    excel_file = xlsx_file

    # Read the Excel file into a Pandas DataFrame
    df = pd.read_excel(excel_file)
    masterdatadf = pd.ExcelFile(r"C:\Users\rpauser\Documents\Item Master Date.xlsx")

    # Define the data you want to add in the new column

    mapped_col=[]
    found_col=[]
    divisionfactor_col=[]

    for index,row in df.iterrows():
        
        if pd.isnull(row[1]):
            mapped_BTC_code=""
            division_factor=""
            FOUND=""            
        else:
            mapped_BTC_code,division_factor,FOUND = mapcustomer_sku(row[1],"Dark Store",masterdatadf)
        #print(mapped_BTC_code)
        mapped_col.append(mapped_BTC_code)
        found_col.append(FOUND)
        divisionfactor_col.append(division_factor)
    #new_column_data = ['NewData1', 'NewData2', 'NewData3', 'NewData4']  # Replace with your data

    # Create a new column and fill it with the new data
    new_column_name = 'MappedSKU'  # Name of the new column
    df[new_column_name] = mapped_col

    found_col_name='FOUND'
    df[found_col_name]=found_col

    div_col_name='DF'
    df[div_col_name]=divisionfactor_col

    # Save the updated DataFrame back to the Excel file
    df.to_excel(excel_file, index=False,sheet_name='Original')

add_mappeddata_toDF(csv_path1)

def extract_store_information(pdf_path):
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        page = reader.pages[0]
        text = page.extract_text()

        # Extract Store Information
        store_info = re.search(r'Store\s*Information:\s*(.*?)(?:\n|$)', text)
        store_info_str = store_info.group(1).strip() if store_info else None

        # Remove all spaces from the store information
        if store_info_str:
            store_info_str = re.sub(r'\s+', '', store_info_str)

        return store_info_str

store_info = extract_store_information(pdf_path1)

print("Store Information:", store_info)



excel_file = pd.ExcelFile(csv_path1)


# Create a new sheet for PO info
with pd.ExcelWriter(csv_path1, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
    # Write the purchase order details to the new sheet
    po_info_df = pd.DataFrame({
        'Store Information': [store_info],
    })
    po_info_df.to_excel(writer, sheet_name='po info', index=False)


print("Purchase order details saved to Excel sheet.")
# Function to copy the first two rows from "Original" sheet to "po info" sheet

# Function to copy the first two rows from "Original" sheet to "po info" sheet and delete them from "Original" sheet
def copy_first_two_rows(xlsx_file):
    # Load the workbook
    wb = openpyxl.load_workbook(xlsx_file)

    # Select the "Original" sheet
    original_sheet = wb["Original"]

    # Get the first two rows from "Original" sheet
    rows_to_copy = list(original_sheet.iter_rows(min_row=1, max_row=2, values_only=True))

    # Create or select "po info" sheet
    if "po info" in wb.sheetnames:
        po_info_sheet = wb["po info"]
    else:
        po_info_sheet = wb.create_sheet("po info")

    # Copy the rows to "po info" sheet
    for row in rows_to_copy:
        po_info_sheet.append(row)

    # Delete the copied rows from "Original" sheet
    for i in range(len(rows_to_copy)):
        original_sheet.delete_rows(original_sheet.min_row)

    # Save the modified workbook
    wb.save(xlsx_file)

    print("Copying rows from 'Original' to 'po info' and deleting them complete!")


# Usage example:
copy_first_two_rows(csv_path1)

