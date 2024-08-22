import re

import PyPDF2
import openpyxl
import pdfplumber
import pandas as pd
from datetime import datetime


# Set the path to your PDF file
pdf_path1 = "2311002875094.pdf"

# Initialize an empty list to store the extracted tables
tables = []

# Initialize pdfplumber object
with pdfplumber.open(pdf_path1) as pdf:
    # Iterate over each page and extract the tables
    for page in pdf.pages:
        # Extract tables from the current page
        page_tables = page.extract_tables()
        # Append the extracted tables to the list
        tables.extend(page_tables)

# Combine the extracted tables into a single DataFrame
combined_table = pd.concat([pd.DataFrame(table) for table in tables], ignore_index=True)

# Set the path to save the CSV file
csv_path1 = "output.xlsx"

# Save the DataFrame as an Excel file (CSV format)
combined_table.to_excel(csv_path1, sheet_name='Sheet1', index=False, header=False)
def delete_rows_from_excel(file_path, sheet_name, first_rows=5, last_rows=6):
    try:
        # Open the Excel file using openpyxl
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        # Check if there are enough rows to delete
        if sheet.max_row >= first_rows + last_rows:
            # Delete the first rows
            for _ in range(first_rows):
                sheet.delete_rows(1)

            # Delete the last rows
            for _ in range(last_rows):
                sheet.delete_rows(sheet.max_row)

            # Save the modified Excel file
            workbook.save(file_path)
            print(f"Deleted {first_rows} rows from the beginning and {last_rows} rows from the end in '{sheet_name}' sheet of '{file_path}'.")

        else:
            print("Not enough rows to delete.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")

# Example usage:
# Specify the file path and sheet name
sheet_name = "Sheet1"
delete_rows_from_excel(csv_path1, sheet_name, first_rows=5, last_rows=6)

def delete_columns_except(filename, sheet_name):
    # Load the Excel file
    workbook = openpyxl.load_workbook(filename)

    # Select the desired worksheet
    sheet = workbook[sheet_name]

    # Define the columns to keep
    columns_to_keep = [1, 3, 12, 13, 14]

    # Create a list of columns to delete
    columns_to_delete = [col for col in range(1, sheet.max_column + 1) if col not in columns_to_keep]

    # Iterate over the columns to delete in reverse order (to avoid shifting column indexes)
    for col in reversed(columns_to_delete):
        sheet.delete_cols(col)

    # Save the modified workbook
    workbook.save(filename)

    print(f"Columns except {columns_to_keep} have been deleted from {sheet_name} sheet in {filename}.")


# Usage example:
worksheet_name = 'Sheet1'  # Change this to the name of your worksheet
delete_columns_except(csv_path1, worksheet_name)
def swap_columns_in_excel(file_path, sheet_name):
    try:
        # Read the Excel file into a DataFrame without headers
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

        # Swap the first and second columns
        if len(df.columns) >= 2:
            df[df.columns[0]], df[df.columns[1]] = df[df.columns[1]].copy(), df[df.columns[0]].copy()

        # Reset the index to start from 0
        df.reset_index(drop=True, inplace=True)

        # Save the modified DataFrame back to the same sheet in the Excel file
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

        print(f"Columns swapped and saved back to sheet '{sheet_name}' in '{file_path}'.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")
sheet_name = "Sheet1"
swap_columns_in_excel(csv_path1, sheet_name)
def delete_first_column(file_path, sheet_name):
    try:
        # Load the Excel file
        wb = openpyxl.load_workbook(file_path)

        # Select the specified sheet
        sheet = wb[sheet_name]

        # Loop through the rows and delete the first cell (column A) in each row
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=1):
            for cell in row:
                cell.value = None  # Set the cell value to None to keep it empty

        # Save the modified Excel file
        wb.save(file_path)

        print(f"First column deleted and kept empty in sheet '{sheet_name}' of '{file_path}'.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")

# Example usage:
sheet_name = "Sheet1"  # Replace with the sheet name
delete_first_column(csv_path1, sheet_name)
def extract_purchase_order_details(pdf_path):
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        # Extract text from the first page
        first_page = reader.pages[0]
        first_page_text = first_page.extract_text()
        # Extract text from the last page
        last_page = reader.pages[-1]
        last_page_text = last_page.extract_text()
        # Combine the text from both pages
        text = first_page_text + "\n" + last_page_text
        print(text)
        # Extract text after "contractAd. ch."
        contract_ad_match = re.search(r'contractAd\. ch\.\s*(.*?)Asked delivery date Delivery deadline', text,
                                      re.DOTALL)
        contract_ad_text = contract_ad_match.group(1) if contract_ad_match else None

        # Extract order number, order date, and supplier code
        order_match = re.search(r'(\d{13}) (\d{2}/\d{2}/\d{2} \d{2}:\d{2}) (\d+) (\S+ \d)', contract_ad_text)
        order_no = order_match.group(1) if order_match else None
        order_date = order_match.group(2) if order_match else None
        supplier_code = order_match.group(3) if order_match else None

        # Extract text after "Asked delivery date Delivery deadline"
        delivery_date_match = re.search(r'Asked delivery date Delivery deadline\s*(.*?)Comment', text, re.DOTALL)
        delivery_date_text = delivery_date_match.group(1) if delivery_date_match else None

        # Extract RDD (Requested Delivery Date) and Expiry
        rdd_match = re.search(r'(\d{2}/\d{2}/\d{2} \d{2}:\d{2}) (\d{2}/\d{2}/\d{2} \d{2}:\d{2})', delivery_date_text)
        rdd = rdd_match.group(1) if rdd_match else None
        expiry = rdd_match.group(2) if rdd_match else None

        return order_no, order_date, supplier_code, rdd, expiry


po_str, po_date,store_info, rdd, expiry = extract_purchase_order_details(pdf_path1)

# Load the Excel file
excel_file = pd.ExcelFile(csv_path1)

def convert_date_format(input_date, input_format, output_format):
    try:
        # Parse the input date string into a datetime object using the input format
        datetime_obj = datetime.strptime(input_date, input_format)

        # Convert the datetime object to the desired output format
        output_date = datetime_obj.strftime(output_format)
        return output_date

    except ValueError:
        return "Invalid input date format"

# Create a new sheet for PO info
with pd.ExcelWriter(csv_path1, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
    # Write the purchase order details to the new sheet
    po_info_df = pd.DataFrame({
        'Purchase Order Number': [po_str],
        'WH Address': [store_info],
        'PO Date': [convert_date_format(po_date.split(" ")[0],'%d/%m/%y','%d.%m.%Y')],
        'Not Before Date': [convert_date_format(rdd.split(" ")[0],'%d/%m/%y','%d.%m.%Y')],
        'Not After Date': [convert_date_format(expiry.split(" ")[0],'%d/%m/%y','%d.%m.%Y')],


    })
    po_info_df.to_excel(writer, sheet_name='po info', index=False)


print("PO No: " + (po_str or "Not Found"))
print("Address: " + (store_info or "Not Found"))
print("PO Date: " + (po_date or "Not Found"))
print("RDD Date: " + (rdd or "Not Found"))
print("Expiry Date: " + (expiry or "Not Found"))
