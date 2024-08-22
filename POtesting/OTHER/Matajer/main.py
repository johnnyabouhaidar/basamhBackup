import re
import PyPDF2
import openpyxl
import pdfplumber
import pandas as pd
import tabula
from datetime import datetime, timedelta

# Set the path to your PDF file
pdf_path1 = "Matajer AlSaudia Com..pdf"

# Extract tables from the PDF using tabula
tables = tabula.read_pdf(pdf_path1, pages='all', multiple_tables=True, encoding='latin')

# Combine the extracted tables into a single DataFrame
combined_table = pd.concat(tables, ignore_index=True)

# Set the path to save the Excel file
excel_path = "output.xlsx"

# Save the DataFrame as an Excel file
with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    combined_table.to_excel(writer, sheet_name='Sheet1', index=False, header=False)

print(f"Tables extracted and saved to '{excel_path}'.")

def delete_columns_except(filename, sheet_name):
    # Load the Excel file
    workbook = openpyxl.load_workbook(filename)

    # Select the desired worksheet
    sheet = workbook[sheet_name]

    # Define the columns to keep
    columns_to_keep = [1,2,3,4,5,6, 7,8, 9,10,11,12,13]

    # Create a list of columns to delete
    columns_to_delete = [col for col in range(1, sheet.max_column + 1) if col not in columns_to_keep]

    # Iterate over the columns to delete in reverse order (to avoid shifting column indexes)
    for col in reversed(columns_to_delete):
        sheet.delete_cols(col)

    # Save the modified workbook
    workbook.save(filename)

    print(f"Columns except {columns_to_keep} have been deleted from {sheet_name} sheet in {filename}.")


# Usage example:
worksheet_name = 'Sheet1'  # Change this to the name of your worksheet
delete_columns_except(excel_path, worksheet_name)


def delete_rows_with_empty_first_column(file_path, sheet_name):
    try:
        # Load the Excel file
        wb = openpyxl.load_workbook(file_path)

        # Select the sheet by name
        sheet = wb[sheet_name]

        # Create a list of rows to delete
        rows_to_delete = []

        # Iterate through the rows
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
            if row[0].value is None:
                rows_to_delete.append(row[0].row)

        # Delete rows in reverse order to avoid shifting issues
        for row_num in reversed(rows_to_delete):
            sheet.delete_rows(row_num)

        # Save the modified workbook back to the same file
        wb.save(file_path)

        print(f"Rows with empty first column deleted in sheet '{sheet_name}' of '{file_path}'.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")


# Example usage:
sheet_name = 'Sheet1'  # Change to the desired sheet name
#delete_rows_with_empty_first_column(excel_path, sheet_name)

def swap_columns_in_excel(file_path, sheet_name):
    try:
        # Read the Excel file into a DataFrame without headers
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

        # Swap the first and second columns
        if len(df.columns) >= 2:
            df[df.columns[0]], df[df.columns[1]] = df[df.columns[1]].copy(), df[df.columns[0]].copy()

        # Reset the index to start from 0
        df.reset_index(drop=True, inplace=True)

        # Save the modified DataFrame back to the same sheet in the Excel file
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

        print(f"Columns swapped and saved back to sheet '{sheet_name}' in '{file_path}'.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")
sheet_name = "Sheet1"
#swap_columns_in_excel(excel_path, sheet_name)


def split_column_3_at_first_space(file_path, sheet_name):
    try:
        # Load the Excel file
        wb = openpyxl.load_workbook(file_path)

        # Select the sheet by name
        sheet = wb[sheet_name]

        # Iterate through the rows
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=3, max_col=3):
            for cell in row:
                # Split the content at the first space
                content = cell.value
                if content and ' ' in content:
                    parts = content.split(' ', 1)
                    # Keep only the second part
                    sheet.cell(row=cell.row, column=3).value = parts[1]

        # Save the modified workbook back to the same file
        wb.save(file_path)

        print(f"Column 3 split at the first space in sheet '{sheet_name}' of '{file_path}'.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")


# Example usage:
sheet_name = 'Sheet1'  # Change to the desired sheet name
#split_column_3_at_first_space(excel_path, sheet_name)


def add_Given_days(nbr_of_days, date_string, format_string):
    # Convert the date string to a datetime object using the provided format
    date = datetime.strptime(date_string, format_string)

    # Add 10 days to the date
    new_date = date + timedelta(days=nbr_of_days)

    # Convert the new date back to a string using the provided format
    new_date_string = new_date.strftime(format_string)

    # Return the new date string
    return new_date_string

def extract_purchase_order_details(pdf_path):
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)

        page = reader.pages[0]
        text = page.extract_text()
        print(text)

        # Extract Purchase Order Number
        po_number = re.search(r'PO Number:\s*([^\n]+)', text)
        po_str = po_number.group(1).strip() if po_number else None

        # Extract Ship To
        contact_match = re.search(r'Contact:\s*([^\n]+)\n([^\n]+)', text)
        ship_to_location_str = contact_match.group(1).strip() if contact_match else None

        # Extract Approval Date and RDD Date
        date_match = re.findall(r'(\d{2}/\d{2}/\d{4})', text)
        if date_match:
            approval_date_str = date_match[0]
            rdd_date_str = date_match[1] if len(date_match) > 1 else "Not Found"
        else:
            approval_date_str = "Not Found"
            rdd_date_str = "Not Found"

        return po_str, ship_to_location_str, approval_date_str, rdd_date_str

po_str, shipto, aprrovalDate, rdd_date = extract_purchase_order_details(pdf_path1)

# Load the Excel file
excel_file = pd.ExcelFile(excel_path)

# Create a new sheet for PO info
with pd.ExcelWriter(excel_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
    # Write the purchase order details to the new sheet
    po_info_df = pd.DataFrame({
        'Purchase Order Number': [po_str],
        'Ship to': [shipto],
        'Order Date': [aprrovalDate],
        'RDD': [add_Given_days(4,aprrovalDate,"%d.%m.%Y")],
        'Expiry Date': [add_Given_days(7,aprrovalDate,"%d.%m.%Y")]
    })
    po_info_df.to_excel(writer, sheet_name='po info', index=False)

print("PO No: " + (po_str or "Not Found"))
print("Ship to: " + (shipto or "Not Found"))
print("Approval Date: " + (aprrovalDate or "Not Found"))
print("RDD Date: " + (rdd_date or "Not Found"))