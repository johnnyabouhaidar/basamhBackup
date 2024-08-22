import re

import PyPDF2
import pandas as pd
import pdfplumber
import openpyxl
# Set the path to your PDF file
pdf_path1 = "C:\\Users\\user\\Desktop\\PDFPOS\\Jumlaty\\Jumlaty.pdf"

# Initialize an empty list to store tables from all pages
all_tables = []

# Determine the total number of pages in the PDF
with pdfplumber.open(pdf_path1) as pdf:
    total_pages = len(pdf.pages)
# Loop through each page and extract tables
for page_num in range(1, total_pages + 1):
    with pdfplumber.open(pdf_path1) as pdf:
        page = pdf.pages[page_num - 1]
        tables = page.extract_tables()

        # Append the tables from the current page to the list
        all_tables.extend(tables)

# Combine the extracted tables into a single DataFrame
combined_table = pd.concat([pd.DataFrame(table) for table in all_tables], ignore_index=True)

# Set the path to save the Excel file
csv_path1 = "C:\\Users\\user\\Desktop\\excelSheetsFolder\\Jumlaty.xlsx"

# Save the DataFrame as an Excel file
combined_table.to_excel(csv_path1, sheet_name='Sheet1', index=False, header=False)

def remove_empty_first_column_cells(filename):
    # Open the Excel file
    wb = openpyxl.load_workbook(filename)

    # Select the first sheet (you can change this if needed)
    sheet = wb.active

    # Get the maximum row count in the sheet
    max_row = sheet.max_row

    # Iterate through rows
    for row_num in range(1, max_row + 1):
        # Get the value of the first cell in the current row
        first_cell_value = sheet.cell(row=row_num, column=1).value

        # Check if the first cell is empty
        if first_cell_value is None:
            # Shift cells 2 to 5 to the left
            for col_num in range(2, 6):
                sheet.cell(row=row_num, column=col_num - 1).value = sheet.cell(row=row_num, column=col_num).value
                # Clear the original cell
                sheet.cell(row=row_num, column=col_num).value = None

    # Save the modified Excel file
    wb.save(filename)
    wb.close()

# Usage:
remove_empty_first_column_cells(csv_path1)


def delete_empty_rows(filename):
    # Open the Excel file
    wb = openpyxl.load_workbook(filename)

    # Select the first sheet (you can change this if needed)
    sheet = wb.active

    # Iterate through rows in reverse order
    for row_num in range(sheet.max_row, 0, -1):
        # Check if all cells in the current row are empty
        if all(sheet.cell(row=row_num, column=col).value is None for col in range(1, sheet.max_column + 1)):
            # Delete the entire row
            sheet.delete_rows(row_num)

    # Save the modified Excel file
    wb.save(filename)
    wb.close()

# Usage:
delete_empty_rows(csv_path1)

def delete_rows_before_keyword(excel_path, sheet_name, keyword):
    # Open the Excel file
    wb = openpyxl.load_workbook(excel_path)

    # Select the sheet by name
    sheet = wb[sheet_name]

    # Find the row index where the keyword is present
    target_row_index = None
    for row_idx, row in enumerate(sheet.iter_rows(min_col=1), start=1):
        if row[0].value == keyword:
            target_row_index = row_idx
            break

    if target_row_index is not None:
        # Delete rows before and including the target row
        sheet.delete_rows(1, target_row_index)

        # Save the modified Excel file
        wb.save(excel_path)
        print("Rows deleted successfully.")
    else:
        print("Keyword not found in the sheet.")

# Example usage
sheet_name = 'Sheet1'
keyword = 'ITEM'
delete_rows_before_keyword(csv_path1, sheet_name, keyword)


def split_and_keep_first_part(filename):
    # Open the Excel file
    wb = openpyxl.load_workbook(filename)

    # Select the first sheet (you can change this if needed)
    sheet = wb.active

    # Iterate through rows
    for row_num in range(1, sheet.max_row + 1):
        # Get the value of the cell in the first column
        cell_value = sheet.cell(row=row_num, column=1).value

        # Check if the cell contains "Name:"
        if cell_value and "Name:" in cell_value:
            # Split the cell value by "Name:" and keep the first part
            first_part = cell_value.split("Name:")[0].strip()

            # Update the cell value with the first part
            sheet.cell(row=row_num, column=1).value = first_part

    # Save the modified Excel file
    wb.save(filename)
    wb.close()
# Usage:
split_and_keep_first_part(csv_path1)


def split_and_keep_second_part(filename):
    # Open the Excel file
    wb = openpyxl.load_workbook(filename)

    # Select the first sheet (you can change this if needed)
    sheet = wb.active

    # Iterate through rows
    for row_num in range(1, sheet.max_row + 1):
        # Get the value of the cell in the first column
        cell_value = sheet.cell(row=row_num, column=1).value

        # Check if the cell value is not empty and contains a space
        if cell_value and ":" in cell_value:
            # Split the cell value by the first space and keep the second part
            second_part = cell_value.split(":", 1)[1].strip()

            # Update the cell value with the second part
            sheet.cell(row=row_num, column=1).value = second_part

    # Save the modified Excel file
    wb.save(filename)
    wb.close()

# Usage:
split_and_keep_second_part(csv_path1)


def split_and_keep_first_part_col3(filename):
    # Open the Excel file
    wb = openpyxl.load_workbook(filename)

    # Select the first sheet (you can change this if needed)
    sheet = wb.active

    # Iterate through rows
    for row_num in range(1, sheet.max_row + 1):
        # Get the value of the cell in the third column
        cell_value = sheet.cell(row=row_num, column=3).value

        # Check if the cell value is not empty and contains a space
        if cell_value and " " in cell_value:
            # Split the cell value by the first space and keep the first part
            first_part = cell_value.split(" ", 1)[0].strip()

            # Update the cell value with the first part
            sheet.cell(row=row_num, column=3).value = first_part

    # Save the modified Excel file
    wb.save(filename)
    wb.close()
# Usage:
split_and_keep_first_part_col3(csv_path1)


def split_and_keep_first_part_col4(filename):
    # Open the Excel file
    wb = openpyxl.load_workbook(filename)

    # Select the first sheet (you can change this if needed)
    sheet = wb.active

    # Iterate through rows
    for row_num in range(1, sheet.max_row + 1):
        # Get the value of the cell in the fourth column
        cell_value = sheet.cell(row=row_num, column=4).value

        # Check if the cell value is not empty and contains a space
        if cell_value and " " in cell_value:
            # Split the cell value by the first space and keep the first part
            first_part = cell_value.split(" ", 1)[0].strip()

            # Update the cell value with the first part
            sheet.cell(row=row_num, column=4).value = first_part

    # Save the modified Excel file
    wb.save(filename)
    wb.close()
# Usage:
split_and_keep_first_part_col4(csv_path1)


def shift_columns_and_leave_col2_empty(filename):
    # Open the Excel file
    wb = openpyxl.load_workbook(filename)

    # Select the first sheet (you can change this if needed)
    sheet = wb.active

    # Iterate through rows
    for row_num in range(1, sheet.max_row + 1):
        # Get the values of columns 2, 3, and 4
        col2_value = sheet.cell(row=row_num, column=2).value
        col3_value = sheet.cell(row=row_num, column=3).value
        col4_value = sheet.cell(row=row_num, column=4).value

        # Shift columns 2, 3, and 4 to the right
        sheet.cell(row=row_num, column=5).value = col4_value
        sheet.cell(row=row_num, column=4).value = col3_value
        sheet.cell(row=row_num, column=3).value = col2_value
        sheet.cell(row=row_num, column=2).value = None  # Make column 2 empty

    # Save the modified Excel file
    wb.save(filename)
    wb.close()

# Usage:
shift_columns_and_leave_col2_empty(csv_path1)


def delete_rows_with_empty_col5(filename):
    # Open the Excel file
    wb = openpyxl.load_workbook(filename)

    # Select the first sheet (you can change this if needed)
    sheet = wb.active

    # Iterate through rows in reverse order
    for row_num in range(sheet.max_row, 0, -1):
        # Get the value of the cell in column 5
        col5_value = sheet.cell(row=row_num, column=5).value

        # Check if the cell in column 5 is empty
        if col5_value is None:
            # Delete the entire row
            sheet.delete_rows(row_num)

    # Save the modified Excel file
    wb.save(filename)
    wb.close()

# Usage:
delete_rows_with_empty_col5(csv_path1)

def extract_purchase_order_details(pdf_path):
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)

        page = reader.pages[0]
        text = page.extract_text()
        # Extract Purchase Order Number
        po_number = re.search(r'(\S+)\n', text)
        po_str = po_number.group(1) if po_number else None

        # Extract Ship To based on "SHIP TO" and the value on the new line after it
        ship_to_match = re.search(r'SHIP\s*TO\s*([\s\S]*?)\n', text)
        ship_to_location_str = ship_to_match.group(1).strip() if ship_to_match else None

        # Extract Order Date
        order_date = re.search(r'\n(\S+)\n', text)
        order_date_str = order_date.group(1) if order_date else None


        return po_str, ship_to_location_str, order_date_str

po_str, shipto, aprrovalDate = extract_purchase_order_details(pdf_path1)

# Load the Excel file
excel_file = pd.ExcelFile(csv_path1)

# Create a new sheet for PO info
with pd.ExcelWriter(csv_path1, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
    # Write the purchase order details to the new sheet
    po_info_df = pd.DataFrame({
        'Purchase Order Number': [po_str],
        'Ship to': [shipto],
        'Order Date': [aprrovalDate],
        'RDD': [''],
        'Expiry Date': ['']

    })
    po_info_df.to_excel(writer, sheet_name='po info', index=False)


print("PO No: " + (po_str or "Not Found"))
print("Ship to: " + (shipto or "Not Found"))
print("Approval Date: " + (aprrovalDate or "Not Found"))
