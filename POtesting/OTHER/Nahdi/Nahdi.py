import PyPDF2
import openpyxl
import pdfplumber
import pandas as pd
import re
from datetime import datetime


# Set the path to your PDF file
pdf_path1 = "Nahdi Com..pdf"
csv_path1 = "output.xlsx"

# Initialize an empty list to store the extracted tables
tables = []

# Initialize pdfplumber object
with pdfplumber.open(pdf_path1) as pdf:
    # Iterate over each page and extract the tables
    for page in pdf.pages:
        # Extract tables from the current page
        page_tables = page.extract_tables()
        # Append the extracted tables to the list
        tables.extend(page_tables)

# Combine the extracted tables into a single DataFrame
combined_table = pd.concat([pd.DataFrame(table) for table in tables], ignore_index=True)

# Set the path to save the CSV file


# Save the DataFrame as an Excel file (CSV format)
combined_table.to_excel(csv_path1, sheet_name='Sheet1', index=False, header=False)
def delete_columns_except(filename, sheet_name):
    # Load the Excel file
    workbook = openpyxl.load_workbook(filename)

    # Select the desired worksheet
    sheet = workbook[sheet_name]

    # Define the columns to keep
    columns_to_keep = [2, 5, 6, 7]

    # Create a list of columns to delete
    columns_to_delete = [col for col in range(1, sheet.max_column + 1) if col not in columns_to_keep]

    # Iterate over the columns to delete in reverse order (to avoid shifting column indexes)
    for col in reversed(columns_to_delete):
        sheet.delete_cols(col)

    # Save the modified workbook
    workbook.save(filename)

    print(f"Columns except {columns_to_keep} have been deleted from {sheet_name} sheet in {filename}.")


# Usage example:
worksheet_name = 'Sheet1'  # Change this to the name of your worksheet
delete_columns_except(csv_path1, worksheet_name)



def delete_rows_with_keyword_or_empty_column(file_path, sheet_name):
    try:
        # Load the Excel workbook
        workbook = openpyxl.load_workbook(file_path)

        # Select the specific sheet by name
        sheet = workbook[sheet_name]

        # Iterate through rows in reverse order to avoid shifting rows
        rows_to_delete = []
        for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row):
            cell_value = row[0].value
            if cell_value is None or "Item No" in str(cell_value):
                rows_to_delete.append(row)

        # Delete rows
        for row in rows_to_delete:
            sheet.delete_rows(row[0].row)

        # Save the modified workbook
        workbook.save(file_path)
        workbook.close()

        print(f"Rows with 'Item No' or empty first column deleted in '{sheet_name}' of '{file_path}'.")

    except Exception as e:
        print(f"An error occurred: {str(e)}")

sheet_name = "Sheet1"
delete_rows_with_keyword_or_empty_column(csv_path1, sheet_name)
def shift_columns_and_leave_col2_empty(filename):
    # Open the Excel file
    wb = openpyxl.load_workbook(filename)

    # Select the first sheet (you can change this if needed)
    sheet = wb.active

    # Iterate through rows
    for row_num in range(1, sheet.max_row + 1):
        # Get the values of columns 2, 3, and 4
        col2_value = sheet.cell(row=row_num, column=2).value
        col3_value = sheet.cell(row=row_num, column=3).value
        col4_value = sheet.cell(row=row_num, column=4).value

        # Shift columns 2, 3, and 4 to the right
        sheet.cell(row=row_num, column=5).value = col4_value
        sheet.cell(row=row_num, column=4).value = col3_value
        sheet.cell(row=row_num, column=3).value = col2_value
        sheet.cell(row=row_num, column=2).value = None  # Make column 2 empty

    # Save the modified Excel file
    wb.save(filename)
    wb.close()

# Usage:
shift_columns_and_leave_col2_empty(csv_path1)

def swap_columns_in_excel(file_path, sheet_name):
    try:
        # Read the Excel file into a DataFrame without headers
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        
        # Swap the first and second columns
        if len(df.columns) >= 2:
            df[df.columns[0]], df[df.columns[1]] = df[df.columns[1]].copy(), df[df.columns[0]].copy()

        # Reset the index to start from 0
        df.reset_index(drop=True, inplace=True)
        print(df)
        df[0]="-"
        df[4]=df[4].str.replace("\n","")

        
        # Save the modified DataFrame back to the same sheet in the Excel file
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

        print(f"Columns swapped and saved back to sheet '{sheet_name}' in '{file_path}'.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")
sheet_name = "Sheet1"
swap_columns_in_excel(csv_path1, sheet_name)

def extract_purchase_order_details(pdf_path):
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)

        page = reader.pages[0]
        text = page.extract_text()
        # Extract Purchase Order Number
        po_number = re.search(r'PO NO.\s*:\s*(\d+)', text)
        po_str = po_number.group(1) if po_number else None
        # Extract Store Information
        store_info = re.search(r'WH Address\s*:\s*(.*?)(?:\n|$)', text)
        store_info_str = store_info.group(1).strip() if store_info else None
        # Extract PO Date
        po_date = re.search(r'PO Date\s*:\s*([-0-9A-Za-z]+)', text)
        po_date_str = po_date.group(1) if po_date else None
        # Extract not before date
        not_before_date = re.search(r'Not Before Date\s*:\s*([-0-9A-Za-z]+)', text)
        not_before_date_str = not_before_date.group(1) if not_before_date else None
        # Extract Not after date
        not_after_date = re.search(r'Not After Date\s*:\s*([-0-9A-Za-z]+)', text)
        not_after_date_str = not_after_date.group(1) if not_after_date else None

        return po_str, store_info_str, po_date_str, not_before_date_str, not_after_date_str


po_str, store_info, po_date, not_before_date, not_after_date = extract_purchase_order_details(pdf_path1)

# Load the Excel file
excel_file = pd.ExcelFile(csv_path1)
def convert_date_format(input_date, input_format, output_format):
    try:
        # Parse the input date string into a datetime object using the input format
        datetime_obj = datetime.strptime(input_date, input_format)

        # Convert the datetime object to the desired output format
        output_date = datetime_obj.strftime(output_format)
        return output_date

    except ValueError:
        return "Invalid input date format"


# Create a new sheet for PO info
with pd.ExcelWriter(csv_path1, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
    # Write the purchase order details to the new sheet
    po_info_df = pd.DataFrame({
        'Purchase Order Number': [po_str],
        'WH Address': [store_info],
        'PO Date': [convert_date_format(po_date,'%d-%b-%Y','%d.%m.%Y')],
        'Not Before Date': [convert_date_format(not_before_date,'%d-%b-%Y','%d.%m.%Y')],
        'Not After Date': [convert_date_format(not_after_date,'%d-%b-%Y','%d.%m.%Y')],


    })
    po_info_df.to_excel(writer, sheet_name='po info', index=False)


print("PO No: " + (po_str or "Not Found"))
print("Address: " + (store_info or "Not Found"))
print("PO Date: " + (po_date or "Not Found"))
print("Not before Date: " + (not_before_date or "Not Found"))
print("Not after Date: " + (not_after_date or "Not Found"))
