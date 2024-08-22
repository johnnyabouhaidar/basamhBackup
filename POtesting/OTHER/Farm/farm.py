import PyPDF2
import openpyxl
import tabula
import pandas as pd
import re
from datetime import datetime

# Set the path to your PDF file
pdf_path1 = "PurchaseOrderDocument(4655).pdf"
csv_path1 = "output.xlsx"

# Extract tables from the PDF file
tables = tabula.read_pdf(pdf_path1, pages='all', multiple_tables=True, encoding='latin')


# Combine the extracted tables into a single DataFrame
combined_table = pd.concat(tables, ignore_index=True)
combined_table.to_excel("ttt.xlsx")

# Set the path to save the Excel file


# Save the DataFrame as an Excel file
combined_table.to_excel(csv_path1, sheet_name='Sheet1', index=False, header=False)

def delete_columns_except(filename, sheet_name):
    # Load the Excel file
    workbook = openpyxl.load_workbook(filename)

    # Select the desired worksheet
    sheet = workbook[sheet_name]

    # Define the columns to keep (2, 4, 6, 7, 8)
    columns_to_keep = [2, 3, 7, 11, 12,5]

    # Create a list of columns to delete
    columns_to_delete = [col for col in range(1, sheet.max_column + 1) if col not in columns_to_keep]

    # Iterate over the columns to delete in reverse order (to avoid shifting column indexes)
    for col in reversed(columns_to_delete):
        sheet.delete_cols(col)

    # Save the modified workbook
    column_to_move = 'C'  # Change this to the column letter you want to move

    max_row = sheet.max_row
    max_column = sheet.max_column

    # Extract the column data to be moved
    extracted_column = []
    for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=sheet[column_to_move + '1'].column, max_col=sheet[column_to_move + '1'].column):
        for cell in row:
            extracted_column.append(cell.value)

    # Delete the original column
    sheet.delete_cols(sheet[column_to_move + '1'].column)

    # Insert the extracted column data into the last position
    for index, value in enumerate(extracted_column, start=1):
        sheet.cell(row=index, column=max_column, value=value)
    workbook.save(filename)

    print(f"Columns except {columns_to_keep} have been deleted from {sheet_name} sheet in {filename}.")


# Usage example:
worksheet_name = 'Sheet1'  # Change this to the name of your worksheet
delete_columns_except(csv_path1, worksheet_name)

def extract_purchase_order_details(pdf_path):
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)

        page = reader.pages[0]
        text = page.extract_text()
        print(text)
        # Extract Purchase Order Number
        po_number = re.search(r'Order#\s*:\s*([^\n]+)', text)
        po_str = po_number.group(1).strip() if po_number else None

        ship_to_location = re.search(r'Deliver To\s*:\s*(.*?)(?:\s*Order Valid Until|$)', text, re.DOTALL)
        ship_to_location_str = ship_to_location.group(1).strip() if ship_to_location else None

        # Extract Order Date
        approval_date = re.search(r'Order\s*Date:\s*([-0-9A-Za-z/]+)', text)
        approval_date_str = approval_date.group(1) if approval_date else None
        # Extract RDD
        rdd_date = re.search(r'Delivery Before:\s*([-0-9A-Za-z/]+)', text)
        rdd_date_str = rdd_date.group(1) if rdd_date else None
        # Extract Expiry
        expiry_date = re.search(r'Order Valid Until:\s*([-0-9A-Za-z/]+)', text)
        expiry_date_str = expiry_date.group(1) if expiry_date else None

        return po_str, ship_to_location_str, approval_date_str, rdd_date_str, expiry_date_str


po_str, shipto, aprrovalDate, rdd_date, expiry_date = extract_purchase_order_details(pdf_path1)

# Load the Excel file
excel_file = pd.ExcelFile(csv_path1)

def convert_date_format(input_date, input_format, output_format):
    try:
        # Parse the input date string into a datetime object using the input format
        datetime_obj = datetime.strptime(input_date, input_format)

        # Convert the datetime object to the desired output format
        output_date = datetime_obj.strftime(output_format)
        return output_date

    except ValueError:
        return "Invalid input date format"

# Create a new sheet for PO info
with pd.ExcelWriter(csv_path1, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
    # Write the purchase order details to the new sheet
    po_info_df = pd.DataFrame({
        'Purchase Order Number': [po_str],
        'Ship to': [shipto],
        'Order Date': [convert_date_format(aprrovalDate,'%d-%b-%Y','%d.%m.%Y')],
        'RDD': [convert_date_format(rdd_date,'%d-%b-%y','%d.%m.%Y')],
        'Expiry Date': [convert_date_format(expiry_date,'%d-%b-%y','%d.%m.%Y')]

    })
    po_info_df.to_excel(writer, sheet_name='po info', index=False)


print("PO No: " + (po_str or "Not Found"))
print("Ship to: " + (shipto or "Not Found"))
print("Approval Date: " + (aprrovalDate or "Not Found"))
print("RDD Date: " + (rdd_date or "Not Found"))
print("Expiry Date: " + (expiry_date or "Not Found"))