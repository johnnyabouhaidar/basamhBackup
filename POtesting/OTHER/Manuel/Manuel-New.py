import PyPDF2
import openpyxl
import pdfplumber
import pandas as pd
import re


# Set the path to your PDF file
pdf_path1 = "C:\\Users\\user\\Desktop\\PDFPOS\\Manuel\\Manuel 2.pdf"

# Initialize an empty list to store the extracted tables
tables = []

# Initialize pdfplumber object
with pdfplumber.open(pdf_path1) as pdf:
    # Iterate over each page and extract the tables
    for page in pdf.pages:
        # Extract tables from the current page
        page_tables = page.extract_tables()
        # Append the extracted tables to the list
        tables.extend(page_tables)


# Combine the extracted tables into a single DataFrame
combined_table = pd.concat([pd.DataFrame(table) for table in tables], ignore_index=True)

# Set the path to save the CSV file
csv_path1 = "C:\\Users\\user\\Desktop\\excelSheetsFolder\\Manuel.xlsx"

# Save the DataFrame as an Excel file (CSV format)
combined_table.to_excel(csv_path1, sheet_name='Sheet1', index=False, header=False)

def process_excel_file(file_path):
    # Load the Excel file
    wb = openpyxl.load_workbook(file_path)

    # Select the original sheet
    original_sheet = wb['Sheet1']

    # Collect rows to delete
    rows_to_delete = []
    for row in original_sheet.iter_rows(min_row=1, min_col=1, max_col=1):
        cell_value = row[0].value
        if any(keyword in cell_value for keyword in ["Order Date:", "N/B Date:", "Expiry Date:", "Created By:"]):
            rows_to_delete.append(row[0].row)

    # Delete the selected rows from the original sheet
    for row_idx in reversed(rows_to_delete):
        original_sheet.delete_rows(row_idx)

    # Save the modified workbook
    wb.save(file_path)

# Call the function to process the Excel file by deleting selected rows from the original sheet
process_excel_file(csv_path1)


def split_excel_data(file_path):
    # Load the Excel file
    wb = openpyxl.load_workbook(file_path)

    # Select the "Original" sheet
    sheet = wb['Sheet1']

    # Iterate through rows and split data
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
        cell = row[0]
        cell_value = cell.value
        if cell_value:
            values = cell_value.split(' ')
            for col_idx, value in enumerate(values, start=1):
                new_cell = sheet.cell(row=cell.row, column=col_idx)
                new_cell.value = value

    # Save the modified workbook
    wb.save(file_path)

# Call the function to split data in the "Original" sheet in-place
split_excel_data(csv_path1)

def delete_rows_before_keyword(excel_path, sheet_name, keyword):
    # Open the Excel file
    wb = openpyxl.load_workbook(excel_path)

    # Select the sheet by name
    sheet = wb[sheet_name]

    # Find the row index where the keyword is present
    target_row_index = None
    for row_idx, row in enumerate(sheet.iter_rows(min_col=1), start=1):
        if row[0].value == keyword:
            target_row_index = row_idx
            break

    if target_row_index is not None:
        # Delete rows before and including the target row
        sheet.delete_rows(1, target_row_index)

        # Save the modified Excel file
        wb.save(excel_path)
        print("Rows deleted successfully.")
    else:
        print("Keyword not found in the sheet.")

# Example usage
sheet_name = 'Sheet1'
keyword = 'Total'
delete_rows_before_keyword(csv_path1, sheet_name, keyword)
def process_column_4(file_path):
    # Load the Excel file
    wb = openpyxl.load_workbook(file_path)

    # Select the "Original" sheet
    sheet = wb['Sheet1']

    # Keep track of rows to shift
    rows_to_shift = []
    old_Value = ''
    # Iterate through cells in column 4 (D)
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=4, max_col=4):
        cell = row[0]
        cell_value = cell.value
        try:
            int(cell_value)
        except:
            old_Value = cell_value
            rows_to_shift.append(cell.row)

    # Shift columns to the right starting from column 4
    for row_idx in reversed(rows_to_shift):
        for col_idx in range(sheet.max_column, 4, -1):
            sheet.cell(row=row_idx, column=col_idx + 1).value = sheet.cell(row=row_idx, column=col_idx).value
        sheet.cell(row=row_idx, column=4).value = None
        sheet.cell(row=row_idx, column=5).value = old_Value

    # Save the modified workbook
    wb.save(file_path)
# Call the function to process column 4 in the "Original" sheet
process_column_4(csv_path1)

def split_newlines_in_column(file_path):
    # Load the Excel file
    wb = openpyxl.load_workbook(file_path)

    # Select the "Original" sheet
    sheet = wb['Sheet1']

    # Keep track of columns to delete
    cols_to_delete = set()

    # Iterate through cells in column 10 (J) and split on newline
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=10, max_col=10):
        cell = row[0]
        cell_value = cell.value
        if cell_value:
            values = cell_value.split('\n')
            for col_idx, value in enumerate(values, start=1):
                new_cell = sheet.cell(row=cell.row, column=cell.column + col_idx)
                new_cell.value = value
            cols_to_delete.add(cell.column)

    # Delete marked columns in reversed order
    for col_idx in reversed(sorted(list(cols_to_delete))):
        sheet.delete_cols(col_idx)

    # Save the modified workbook
    wb.save(file_path)


# Call the function to split newlines in column 10 of the "Original" sheet
split_newlines_in_column(csv_path1)


def delete_columns_except(filename, sheet_name):
    # Load the Excel file
    workbook = openpyxl.load_workbook(filename)

    # Select the desired worksheet
    sheet = workbook[sheet_name]

    # Define the columns to keep (2, 4, 6, 7, 8)
    columns_to_keep = [2, 4, 6, 7, 8]

    # Create a list of columns to delete
    columns_to_delete = [col for col in range(1, sheet.max_column + 1) if col not in columns_to_keep]

    # Iterate over the columns to delete in reverse order (to avoid shifting column indexes)
    for col in reversed(columns_to_delete):
        sheet.delete_cols(col)

    # Save the modified workbook
    workbook.save(filename)

    print(f"Columns except {columns_to_keep} have been deleted from {sheet_name} sheet in {filename}.")


# Usage example:
worksheet_name = 'Sheet1'  # Change this to the name of your worksheet
delete_columns_except(csv_path1, worksheet_name)

def extract_purchase_order_details(pdf_path):
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)

        page = reader.pages[0]
        text = page.extract_text()
        print(text)
        # Extract Purchase Order Number
        po_number = re.search(r'PO\s*:\s*([^\n]+)', text)
        po_str = po_number.group(1).strip() if po_number else None

        # Extract Ship To Location
        ship_to_location = re.search(r'Location\s*:\s*([^\n]+)', text)
        ship_to_location_str = ship_to_location.group(1).strip() if ship_to_location else None

        # Extract Order Date
        approval_date = re.search(r'Order\s*Date:\s*([-0-9A-Za-z/]+)', text)
        approval_date_str = approval_date.group(1) if approval_date else None
        # Extract RDD
        rdd_date = re.search(r'N/B Date:\s*([-0-9A-Za-z/]+)', text)
        rdd_date_str = rdd_date.group(1) if rdd_date else None
        # Extract Expiry
        expiry_date = re.search(r'Expiry Date:\s*([-0-9A-Za-z/]+)', text)
        expiry_date_str = expiry_date.group(1) if expiry_date else None

        return po_str, ship_to_location_str, approval_date_str, rdd_date_str, expiry_date_str


po_str, shipto, aprrovalDate, rdd_date, expiry_date = extract_purchase_order_details(pdf_path1)

# Load the Excel file
excel_file = pd.ExcelFile(csv_path1)

# Create a new sheet for PO info
with pd.ExcelWriter(csv_path1, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
    # Write the purchase order details to the new sheet
    po_info_df = pd.DataFrame({
        'Purchase Order Number': [po_str],
        'Ship to': [shipto],
        'Order Date': [aprrovalDate],
        'RDD': [rdd_date],
        'Expiry Date': [expiry_date]

    })
    po_info_df.to_excel(writer, sheet_name='po info', index=False)


print("PO No: " + (po_str or "Not Found"))
print("Ship to: " + (shipto or "Not Found"))
print("Approval Date: " + (aprrovalDate or "Not Found"))
print("RDD Date: " + (rdd_date or "Not Found"))
print("Expiry Date: " + (expiry_date or "Not Found"))