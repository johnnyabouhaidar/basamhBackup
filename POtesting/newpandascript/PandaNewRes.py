import re
import PyPDF2
import openpyxl
import pdfplumber
import pandas as pd

# Set the path to your PDF file
pdf_path1 = r"PO-57937581.pdf"

# Initialize an empty list to store the extracted tables
tables = []

# Initialize pdfplumber object
with pdfplumber.open(pdf_path1) as pdf:
    # Iterate over each page and extract the tables
    for page in pdf.pages:
        # Extract tables from the current page
        page_tables = page.extract_tables()
        # Append the extracted tables to the list
        tables.extend(page_tables)

# Combine the extracted tables into a single DataFrame
combined_table = pd.concat([pd.DataFrame(table) for table in tables], ignore_index=True)

# Set the path to save the CSV file
csv_path1 = "Panda33.xlsx"

# Save the DataFrame as an Excel file (CSV format)
combined_table.to_excel(csv_path1, sheet_name='Original', index=False, header=False)


def extract_purchase_order_details(pdf_path):
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)

        page = reader.pages[0]
        text = page.extract_text()
        # Extract Purchase Order Number
        po_number = re.search(r'Purchase Order\s*-\s*(\d+)', text)
        po_str = po_number.group(1) if po_number else None

        # Extract Address
        address = re.search(r'Location\s*Name\s*([\s\S]*?)\s*Supplier Site', text)
        address_str = address.group(1).strip() if address else None

        # Extract Not Before Date
        not_before_date = re.search(r'Not Before Date\s*([-0-9A-Za-z]+)', text)
        not_before_str = not_before_date.group(1) if not_before_date else None

        # Extract Not After Date
        not_after_date = re.search(r'Not After Date\s*([-0-9A-Za-z]+)', text)
        not_after_str = not_after_date.group(1) if not_after_date else None

        # PO context
        PO_Context = re.search(r'PO Context\s*([-0-9A-Za-z]+)', text)
        PO_Context_str = PO_Context.group(1) if PO_Context else None

        # PO creation
        PO_Creation = re.search(r'PO Creation Date\s*([-0-9A-Za-z]+)', text)
        PO_Creation_str = PO_Creation.group(1) if PO_Creation else None
        return po_str, address_str, not_before_str, not_after_str, PO_Context_str,PO_Creation_str


po_number, address, not_before_date, not_after_date, PO_Context,PO_Creation = extract_purchase_order_details(pdf_path1)
# Load the Excel file
excel_file = pd.ExcelFile(csv_path1)

# Create a new sheet for PO info
with pd.ExcelWriter(csv_path1, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
    # Write the purchase order details to the new sheet
    po_info_df = pd.DataFrame({
        'Purchase Order Number': [po_number],
        'Address': [address],
        'Not Before Date': [not_before_date],
        'Not After Date': [not_after_date],
        'PO Context': [PO_Context],
        'PO Creation':[PO_Creation]
    })
    po_info_df.to_excel(writer, sheet_name='po info', index=False)

if po_number:
    print(po_number)
else:
    print("Purchase Order Number not found.")

if address:
    print("Address:", address)
else:
    print("Address not found.")

if not_before_date:
    print("Not Before Date:", not_before_date)
else:
    print("Not Before Date not found.")

if not_after_date:
    print("Not After Date:", not_after_date)
else:
    print("Not After Date not found.")
if PO_Context:
    print("PO Context:", PO_Context)
else:
    print("PO Context not found.")


def delete_columns_from_excel(file_path):
    # Load the Excel file
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook["Original"]  # Assuming the sheet name is "Sheet1"

    columns_to_delete = [1, 2, 5, 6, 9, 12, 13, 16, 19, 21]

    # Deleting columns in reverse order to avoid shifting issues
    for col_idx in sorted(columns_to_delete, reverse=True):
        sheet.delete_cols(col_idx)

    # Save the modified workbook
    workbook.save(file_path)
    print("Columns deleted and file saved.")


# Provide the path to your Excel file

delete_columns_from_excel(csv_path1)


def delete_empty_rows_from_excel(file_path):
    # Load the Excel file
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook["Original"]

    rows_to_delete = []

    # Identify rows to delete based on empty first and fifth cells
    for row_idx in range(1, sheet.max_row + 1):
        first_cell_value = sheet.cell(row=row_idx, column=1).value
        fifth_cell_value = sheet.cell(row=row_idx, column=5).value

        if first_cell_value is None or fifth_cell_value is None:
            rows_to_delete.append(row_idx)

    # Deleting rows in reverse order to avoid shifting issues
    for row_idx in sorted(rows_to_delete, reverse=True):
        sheet.delete_rows(row_idx)

    # Save the modified workbook
    workbook.save(file_path)
    print("Empty rows deleted and file saved.")


# Provide the path to your Excel file
delete_empty_rows_from_excel(csv_path1)


def delete_rows_with_sku_no(file_path):
    # Load the Excel file
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook["Original"]

    rows_to_delete = []

    # Identify rows to delete
    for row in sheet.iter_rows(min_row=1, max_col=1):
        if row[0].value and str(row[0].value).startswith("SKU No."):
            rows_to_delete.append(row[0].row)

    # Delete identified rows in reverse order to avoid index shifting
    for row_idx in reversed(rows_to_delete):
        sheet.delete_rows(row_idx)

    # Save the modified workbook
    workbook.save(file_path)


delete_rows_with_sku_no(csv_path1)